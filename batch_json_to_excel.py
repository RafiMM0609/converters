import pandas as pd
import json
import os
import sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
import glob

def convert_json_to_excel(json_file, output_folder):
    """
    Convert a single JSON file to Excel format
    """
    try:
        # Read JSON file with UTF-8 encoding to handle special characters
        with open(json_file, 'r', encoding='utf-8') as file:
            data = json.load(file)
        
        # Check if data is a list or dict
        if isinstance(data, list):
            # If it's a list, use it directly
            df = pd.DataFrame(data)
        elif isinstance(data, dict):
            # If it's a dict, get the first key's value
            query_key = list(data.keys())[0]
            df = pd.DataFrame(data[query_key])
        else:
            raise ValueError("JSON data must be either a list or a dictionary")
        
        # Create output filename
        base_name = os.path.basename(json_file)
        excel_name = base_name.replace('.json', '.xlsx')
        excel_file = os.path.join(output_folder, excel_name)
        
        # Export to Excel
        df.to_excel(excel_file, index=False, sheet_name='Data')
        
        # Load workbook for formatting
        wb = load_workbook(excel_file)
        ws = wb['Data']
        
        # Define header color (10b4b1) - convert to RGB format
        header_fill = PatternFill(start_color='FF10B4B1',
                                end_color='FF10B4B1',
                                fill_type='solid')
        
        # Format headers
        for cell in ws[1]:
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', 
                                     vertical='center',
                                     wrap_text=True)
        
        # Freeze pane below header
        ws.freeze_panes = 'A2'
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set column width with some padding
            adjusted_width = min(max_length + 2, 50)  # Max width of 50
            ws.column_dimensions[column_letter].width = adjusted_width
            
            # Set alignment for all cells in column
            for cell in column:
                cell.alignment = Alignment(vertical='center',
                                        wrap_text=True)
        
        # Save the formatted workbook
        wb.save(excel_file)
        print(f"✓ Converted: {base_name} -> {excel_name}")
        return True
        
    except Exception as e:
        print(f"✗ Error converting {json_file}: {e}")
        return False

def batch_convert_json_to_excel():
    """
    Convert all JSON files in jsonuser folder to Excel files in data-excel folder
    """
    # Define paths
    json_folder = 'jsonuser'
    output_folder = 'data-excel'
    
    # Check if jsonuser folder exists
    if not os.path.exists(json_folder):
        print(f"Error: Folder '{json_folder}' not found!")
        return
    
    # Create output folder if it doesn't exist
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
        print(f"Created output folder: {output_folder}")
    
    # Find all JSON files in jsonuser folder
    json_pattern = os.path.join(json_folder, '*.json')
    json_files = glob.glob(json_pattern)
    
    if not json_files:
        print(f"No JSON files found in '{json_folder}' folder!")
        return
    
    print(f"Found {len(json_files)} JSON files to convert...")
    print("-" * 50)
    
    # Convert each JSON file
    success_count = 0
    failed_count = 0
    
    for json_file in json_files:
        if convert_json_to_excel(json_file, output_folder):
            success_count += 1
        else:
            failed_count += 1
    
    # Print summary
    print("-" * 50)
    print(f"Conversion Summary:")
    print(f"✓ Successfully converted: {success_count} files")
    print(f"✗ Failed to convert: {failed_count} files")
    print(f"📁 Excel files saved in: {output_folder}")

if __name__ == "__main__":
    print("Batch JSON to Excel Converter")
    print("=" * 50)
    batch_convert_json_to_excel()
