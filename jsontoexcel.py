import pandas as pd
import json
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter

def convert_json_to_excel():
    try:
        json_file = 'client_outlet_202506031523.json'
        if not os.path.exists(json_file):
            raise FileNotFoundError(f"JSON file '{json_file}' not found")        # Read JSON file with UTF-8 encoding to handle special characters
        with open(json_file, 'r', encoding='utf-8') as file:
            data = json.load(file)
        
        # Check if data is a list or dict
        if isinstance(data, list):
            # If it's a list, use it directly
            df = pd.DataFrame(data)
        elif isinstance(data, dict):
            # If it's a dict, get the first key's value
            query_key = list(data.keys())[0]
            df = pd.DataFrame(data[query_key])
        else:
            raise ValueError("JSON data must be either a list or a dictionary")
        
        # Export to Excel
        excel_file = 'excel-baru.xlsx'
        df.to_excel(excel_file, index=False, sheet_name='Outlets')
        
        # Load workbook for formatting
        wb = load_workbook(excel_file)
        ws = wb['Outlets']
        
        # Define header color (10b4b1) - convert to RGB format
        header_fill = PatternFill(start_color='FF10B4B1',
                                end_color='FF10B4B1',
                                fill_type='solid')
        
        # Format headers
        for cell in ws[1]:
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', 
                                     vertical='center',
                                     wrap_text=True)
        
        # Freeze pane below header
        ws.freeze_panes = 'A2'
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set column width with some padding
            adjusted_width = min(max_length + 2, 50)  # Max width of 50
            ws.column_dimensions[column_letter].width = adjusted_width
            
            # Set alignment for all cells in column
            for cell in column:
                cell.alignment = Alignment(vertical='center',
                                        wrap_text=True)
        
        # Save the formatted workbook
        wb.save(excel_file)
        print(f"Excel file '{excel_file}' has been created successfully!")
        
    except FileNotFoundError as e:
        print(f"Error: {e}")
    except KeyError as e:
        print(f"Error in JSON data: {e}")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

if __name__ == "__main__":
    convert_json_to_excel()
