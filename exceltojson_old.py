import pandas as pd
import json
import sys
from pathlib import Path
from openpyxl import load_workbook

def rgb_to_color_name_modern(r, g, b):
    """Modern pattern matching approach"""
    match (r > 180, g > 180, b > 180):
        case (True, True, True): return "putih"
        case (False, False, False) if all(x < 60 for x in [r,g,b]): return "hitam"
        case (True, False, False): return "merah"
        case (False, True, False): return "hijau" 
        case (False, False, True): return "biru"
        case (True, True, False): return "kuning"
        case _: return "lainnya"

def rgb_to_color_name(rgb_value):
    """Convert RGB to simple color name"""
    if not rgb_value or rgb_value in ["00000000", "FFFFFFFF", "FFFFFF", "000000"]:
        return None
    
    try:
        # Clean up RGB value - handle various formats
        clean_rgb = str(rgb_value).upper()
        
        # Remove any non-hex characters
        clean_rgb = ''.join(c for c in clean_rgb if c in '0123456789ABCDEF')
        
        if len(clean_rgb) == 8:
            # Remove alpha channel (FF prefix/suffix)
            clean_rgb = clean_rgb[-6:] if clean_rgb.startswith('FF') else clean_rgb[:6]
        elif len(clean_rgb) == 6:
            # Already in correct format
            pass
        else:
            return None
        
        # Convert to integers
        r = int(clean_rgb[0:2], 16)
        g = int(clean_rgb[2:4], 16) 
        b = int(clean_rgb[4:6], 16)
        
        # Use modern pattern matching approach
        return rgb_to_color_name_modern(r, g, b)
            
    except (ValueError, TypeError) as e:
        print(f"Debug: Error parsing RGB {rgb_value}: {e}")
        return None

def get_cell_color(cell):
    """Extract color from cell - simplified approach"""
    try:
        # Check fill color first
        if cell.fill and cell.fill.start_color:
            color_obj = cell.fill.start_color
            
            # Try RGB first
            if hasattr(color_obj, 'rgb') and color_obj.rgb:
                rgb_val = color_obj.rgb
                # Handle bytes/string conversion safely
                if isinstance(rgb_val, bytes):
                    return rgb_val.hex().upper()
                else:
                    return str(rgb_val).upper()
            
            # Handle indexed colors with basic mapping
            elif hasattr(color_obj, 'indexed') and color_obj.indexed is not None:
                indexed_map = {
                    2: "FF0000",   # Red
                    3: "00FF00",   # Green
                    4: "0000FF",   # Blue
                    5: "FFFF00",   # Yellow
                    6: "FF00FF",   # Magenta
                    7: "00FFFF",   # Cyan
                    8: "000000",   # Black
                    9: "FFFFFF",   # White
                    10: "800000",  # Dark Red
                    11: "008000",  # Dark Green
                    12: "000080",  # Dark Blue
                    13: "808000",  # Olive
                    14: "800080",  # Purple
                    15: "008080",  # Teal
                    16: "C0C0C0",  # Silver
                    17: "808080",  # Gray
                }
                return indexed_map.get(color_obj.indexed, None)
            
            # Handle theme colors
            elif hasattr(color_obj, 'theme') and color_obj.theme is not None:
                theme_map = {
                    0: "000000",   # Black
                    1: "FFFFFF",   # White
                    2: "1F497D",   # Dark Blue
                    3: "EEECE1",   # Light Gray
                    4: "4F81BD",   # Blue
                    5: "F79646",   # Orange
                    6: "9BBB59",   # Green
                    7: "8064A2",   # Purple
                    8: "4BACC6",   # Light Blue
                    9: "F8696B",   # Light Red
                }
                return theme_map.get(color_obj.theme, None)
                
    except Exception as e:
        print(f"Debug: Error getting cell color: {e}")
        pass
    
    return None

def excel_to_json(excel_file: str, json_file: str = None) -> None:
    """Convert Excel to JSON with color detection from first column of each row"""
    try:
        if not Path(excel_file).exists():
            print(f"Error: File {excel_file} not found!")
            return
        
        print(f"Reading {excel_file}...")
        df = pd.read_excel(excel_file)
        workbook = load_workbook(excel_file)
        worksheet = workbook.active
        
        records = df.to_dict('records')
        
        # Process each row
        for i, record in enumerate(records):
            row_index = i + 2  # Skip header
            
            # Check color of the first column cell only
            first_cell = worksheet.cell(row=row_index, column=1)
            first_cell_color = get_cell_color(first_cell)
            record['color'] = None
            # If first column has color, add to record
            if first_cell_color:
                color_name = rgb_to_color_name(first_cell_color)
                if color_name:
                    # record['warna'] = color_name
                    record['color'] = color_name  # Add color key with same value
            else:
                # If no color detected from cell, check if warna exists and copy to color
                if 'warna' in record and record['warna']:
                    record['color'] = record['warna']
                else:
                    record['color'] = None  # Set to None if no color info available
        
        # Save to JSON
        if json_file is None:
            json_file = Path(excel_file).stem + '_colored.json'
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(records, f, indent=2, ensure_ascii=False)
        
        colored_count = sum(1 for r in records if 'color' in r and r['color'])
        print(f"✅ Converted to {json_file}")
        print(f"📊 {colored_count}/{len(records)} records have color info from first column")
        
    except Exception as e:
        print(f"❌ Error: {str(e)}")

def main():
    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
        json_file = sys.argv[2] if len(sys.argv) > 2 else None
        excel_to_json(excel_file, json_file)
    else:
        print("Usage: python exceltojson.py <excel_file> [json_file]")

if __name__ == "__main__":
    main()